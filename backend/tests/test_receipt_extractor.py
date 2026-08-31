from io import BytesIO

from openpyxl import load_workbook

from backend.app.services.receipt_extractor import (
    build_payment_rows,
    build_row,
    clean_columns,
    classify_expense,
    create_reimbursement_workbook,
    create_workbook,
    extract_invoice_fields,
    extract_known_fields,
)


SAMPLE_LINES = [
    "账单",
    "孝昌那些年餐饮服务管理有限公司",
    "-405.00",
    "支付时间",
    "2026年8月23日 19:36:37",
    "商品",
    "孝昌那些年餐饮服务管理有限公司-消费",
    "商户全称",
    "孝昌那些年餐饮服务管理有限公司",
    "交易单号",
    "4200003244202608239077856598",
]


def test_extract_known_fields_from_receipt_lines():
    fields = extract_known_fields(SAMPLE_LINES)

    assert fields["payment_amount"] == "405.00"
    assert fields["payment_time"] == "2026-08-23 19:36:37"
    assert fields["merchant_name"] == "孝昌那些年餐饮服务管理有限公司"
    assert fields["product_name"].endswith("-消费")
    assert fields["transaction_number"] == "4200003244202608239077856598"


def test_extract_known_fields_from_personal_qr_payment():
    fields = extract_known_fields([
        "账单", "扫一扫付款-给福润烟酒店", "-800.00", "转账时间", "2026年8月22日 17:53:15", "转账单号", "100010730120260822006103709", "53101",
    ])

    assert fields == {
        "payment_amount": "800.00",
        "payment_time": "2026-08-22 17:53:15",
        "merchant_name": "福润烟酒店",
        "product_name": "扫一扫付款",
        "transaction_number": "100010730120260822006103709",
    }


def test_extracts_unlabelled_merchant_title_from_wallet_receipt():
    fields = extract_known_fields([
        "账单", "柒柒小厨餐饮", "-12.00", "当前状态", "支付成功", "收单机构", "财付通支付科技有限公司", "支付时间", "2026年8月5日 19:40:34",
    ])

    assert fields["merchant_name"] == "柒柒小厨餐饮"
    assert fields["product_name"] == "柒柒小厨餐饮"


def test_extract_invoice_uses_tax_inclusive_total_from_digital_invoice():
    fields = extract_invoice_fields([
        "电子发票（增值税专用发票）", "发票号码：", "26424000000104330311", "开票日期：", "2026年08月09日",
        "金额 税额", "￥2376.24 ￥23.76", "价税合计（大写） （小写）", "贰仟肆佰圆整 ￥2400.00",
    ])

    assert fields == {"invoice_amount": "2400.00", "invoice_number": "26424000000104330311", "invoice_date": "2026-08-09"}


def test_build_row_only_includes_requested_columns():
    row = build_row(["付款金额", "付款时间", "商家名称", "备注", "不存在的字段"], SAMPLE_LINES, "receipt.jpg")

    assert row == {
        "源文件": "receipt.jpg",
        "付款金额": "405.00",
        "付款时间": "2026-08-23 19:36:37",
        "商家名称": "孝昌那些年餐饮服务管理有限公司",
        "备注": "孝昌那些年餐饮服务管理有限公司-消费",
        "不存在的字段": "",
    }


def test_create_workbook_uses_requested_headers_and_rows():
    content = create_workbook(["付款金额"], [{"源文件": "a.jpg", "付款金额": "405.00"}])
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    assert list(sheet.values) == [("源文件", "付款金额"), ("a.jpg", "405.00")]


def test_build_payment_rows_adds_invoice_status_and_matches_amount():
    columns, rows = build_payment_rows(
        ["付款金额", "发票金额", "发票号码"],
        [
            ("payment.jpg", SAMPLE_LINES),
            ("invoice.jpg", ["电子发票", "价税合计", "405.00", "发票号码", "INV123456"]),
            ("unmatched.jpg", ["电子发票", "价税合计", "99.00", "发票号码", "INV999999"]),
        ],
    )

    assert columns == ["付款金额", "发票金额", "发票号码", "是否有发票"]
    assert rows == [{
        "源文件": "payment.jpg",
        "付款金额": "405.00",
        "发票金额": "405.00",
        "发票号码": "INV123456",
        "是否有发票": "有（金额匹配）",
        "_invoice_source": "invoice.jpg",
    }]


def test_clean_columns_splits_chinese_enumeration_commas():
    assert clean_columns(["付款金额、付款时间、商家名称、备注"]) == ["付款金额", "付款时间", "商家名称", "备注"]


def test_reimbursement_export_is_compact_and_uses_invoice_period():
    content = create_reimbursement_workbook(
        [{
            "源文件": "payment.jpg",
            "付款金额": "405.00",
            "付款时间": "2026-05-31 19:36:37",
            "商家名称": "某某餐饮服务有限公司",
            "备注": "晚餐消费",
            "发票金额": "405.00",
            "发票号码": "INV123456",
            "发票日期": "2026-06-30",
            "是否有发票": "有（金额匹配）",
            "_invoice_source": "invoice.jpg",
        }],
        {},
        {},
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    payment_sheet = workbook["支付明细"]
    invoice_sheet = workbook["发票明细"]

    assert workbook.sheetnames == ["支付明细", "发票明细"]
    assert payment_sheet["A1"].value.endswith("（2026.06.30-2026.06.30）")
    assert list(payment_sheet.values)[2][:7] == (1, "2026-05-31", "19:36:37", "某某餐饮服务有限公司", 405.0, "餐费", "有票")
    assert payment_sheet["A4"].value == "合计"
    assert payment_sheet["E4"].value == "=SUM(E3:E3)"
    assert payment_sheet.max_row == 4
    assert list(invoice_sheet.values)[1][:5] == (1, "405.00", "405.00", "INV123456", "金额匹配")
    assert invoice_sheet.max_row == 2


def test_expense_classification_has_requested_categories():
    assert classify_expense({"备注": "酒店住宿"}) == "住宿费"
    assert classify_expense({"商家名称": "五金工具店"}) == "工具/材料费"
    assert classify_expense({"商家名称": "滴滴出行"}) == "交通费"
    assert classify_expense({"商家名称": "某餐饮店"}) == "餐费"
    assert classify_expense({"商家名称": "未知商户"}) == "其他"
