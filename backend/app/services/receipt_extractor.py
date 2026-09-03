from __future__ import annotations

import re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from collections.abc import Callable, Sequence
from difflib import SequenceMatcher
from io import BytesIO
from os import getenv
from pathlib import Path
from queue import Queue

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Border, Font, PatternFill, Side


OCRLines = Sequence[str]
OcrReader = Callable[[bytes], OCRLines]
AMOUNT_PATTERN = r"\d+(?:,\d{3})*\.\d{2}"


def normalize_column_name(value: str) -> str:
    return re.sub(r"[\s_\-()（）:：]", "", value).lower()


def clean_columns(raw_columns: Sequence[str]) -> list[str]:
    # Accept natural Chinese field lists pasted with commas, enumeration commas, or newlines.
    columns = [part.strip() for column in raw_columns for part in re.split(r"[，,、\n]+", column) if part.strip()]
    if not columns:
        raise ValueError("请至少填写一列。")
    if len(columns) > 20:
        raise ValueError("首版最多支持 20 列。")
    if len(set(columns)) != len(columns):
        raise ValueError("列名不能重复。")
    return columns


def _next_value(lines: OCRLines, labels: set[str]) -> str:
    normalized_labels = {normalize_column_name(label) for label in labels}
    for index, line in enumerate(lines[:-1]):
        if re.search(r":|\uFF1A", line):
            left, right = re.split(r":|\uFF1A", line, maxsplit=1)
            if normalize_column_name(left) in normalized_labels and right.strip():
                return right.strip()
        if normalize_column_name(line) in normalized_labels:
            for candidate in lines[index + 1 :]:
                value = candidate.strip()
                if value and normalize_column_name(value) not in normalized_labels:
                    return value
    return ""


def _invoice_seller(lines: OCRLines) -> str:
    """Extract the seller name without mistaking the generic invoice label '名称'.

    Invoice OCR commonly splits ``销售方信息 / 名称 / xxx`` into separate lines,
    and may emit a second ``名称：`` for the buyer.  Prefer the value after the
    seller section and reject labels/placeholders.
    """
    cleaned = [re.sub(r"\s+", "", line) for line in lines if line and line.strip()]
    # Some printed invoices are returned by OCR as separate tokens: 名 / 称： / value.
    merged: list[str] = []
    i = 0
    while i < len(cleaned):
        if cleaned[i] in {"名", "名称", "称"} and i + 1 < len(cleaned) and cleaned[i + 1].startswith("称"):
            rest = cleaned[i + 1][1:].lstrip("：:")
            if rest:
                merged.append("名称：" + rest); i += 2
            elif i + 2 < len(cleaned):
                merged.append("名称：" + cleaned[i + 2].lstrip("：:")); i += 3
            else:
                i += 2
            continue
        merged.append(cleaned[i]); i += 1
    cleaned = merged
    seller_start = next((i for i, line in enumerate(cleaned) if "销售方信息" in line or "销货方信息" in line), -1)
    search = cleaned[seller_start + 1:] if seller_start >= 0 else cleaned
    labels = {"名称", "纳税人识别号", "地址电话", "开户行及账号", "购买方", "销售方", "销货方"}
    candidates: list[str] = []
    explicit_names: list[str] = []
    stop_words = ("地址", "电话", "开户", "银行", "账号", "统一社会信用", "纳税人识别", "备注", "开票人", "收款人", "复核", "发票专用章", "发票查验", "登录国家税务", "我要查询", "首页", "http", "https", "税务局网站", "价税合计", "小写", "合计", "更多服务", "更多发票", "发送到邮箱", "打印二维码", "下载到手机", "发票管家", "chinatax", "dppt", "nnfp", "可摇奖城市及活动时间", "圆整")
    for line in search:
        if "发票查验" in line:
            break
        value = re.sub(r"^(名称|销售方|销货方)[:：]?", "", line).strip()
        if not value or value in labels or value.startswith("名称"):
            continue
        if any(word in value for word in stop_words):
            continue
        if any(token in value.lower() for token in ("http", ".com", "税务局", "查验")) or "*" in value or "%" in value or re.search(r"[￥¥]\s*\d", value):
            continue
        if len(re.findall(r"\d", value)) >= 6:
            continue
        if len(re.findall(r"[\u4e00-\u9fffA-Za-z]", value)) >= 2:
            candidates.append(value)
            if re.match(r"^(名称|销售方|销货方)[:：]", line):
                explicit_names.append(value)
    # OCR reading order often emits buyer first and seller second.  The last
    # business-name candidate before address/bank details is therefore the
    # most reliable seller value for common electronic invoices.
    if explicit_names:
        return explicit_names[-1]
    business_words = ("公司", "酒店", "餐馆", "饭店", "烟酒", "商贸", "超市", "餐饮", "食府", "民宿", "酒行")
    business = [value for value in candidates if any(word in value for word in business_words)]
    return (business[-1] if business else candidates[-1]) if candidates else ""


def _usable_invoice_merchant(value: str) -> str:
    """Return a merchant candidate only when it looks like a business name."""
    compact = re.sub(r"\s+", "", value or "")
    if not compact or any(token in compact.lower() for token in ("http", ".com", "税务局", "查验", "网址", "chinatax", "dppt", "nnfp")):
        return ""
    if re.search(r"[￥¥]\s*\d", compact) or re.fullmatch(r"[\d.%，]+", compact):
        return ""
    # Item descriptions and tax breakdowns are not seller names.  They often
    # contain stars, percentages, or a dense run of digits from OCR.
    if "*" in compact or "%" in compact or len(re.findall(r"\d", compact)) >= 6:
        return ""
    if any(token in compact for token in ("价税合计", "小写", "金额", "税额", "开票人", "收款人", "复核", "更多服务", "更多发票", "发送到邮箱", "打印二维码", "下载到手机", "发票管家")):
        return ""
    return compact


def _payment_time(text: str) -> str:
    normalized = re.sub(r"[年月日./]", "-", text)
    normalized = re.sub(r"-(?=\s*\d{1,2}:\d{2})", " ", normalized)
    matched = re.search(r"(20\d{2})\s*-\s*(\d{1,2})\s*-\s*(\d{1,2})(?:\s+|[Tt])*(\d{1,2}:\d{2}(?::\d{2})?)?", normalized)
    if not matched:
        return ""
    year, month, day, clock = matched.groups()
    return f"{year}-{int(month):02d}-{int(day):02d}{f' {clock}' if clock else ''}"


def _personal_payment_title(lines: Sequence[str]) -> tuple[str, str]:
    """Read the payee shown in the large title of personal QR/transfer receipts."""
    for line in lines[:8]:
        compact = re.sub(r"\s+", "", line)
        matched = re.search(r"(?:扫.*?(?:付款|转账)|(?:付款|转账)).*?(?:给|向)(.+)", compact)
        if not matched:
            continue
        payee = matched.group(1).strip("-—－:：")
        description = re.sub(r"[-—－]?(?:给|向).*$", "", compact).strip("-—－:：")
        return payee, description or "个人付款"
    return "", ""


def _unlabelled_merchant_title(lines: Sequence[str]) -> str:
    """Some wallet receipts put the merchant only in the large page title."""
    ignored = ("账单", "支付成功", "收单机构", "支付方式", "支付时间", "付款时间", "转账时间", "交易单号", "经营单号", "当前状态", "账单服务", "收款方")
    for line in lines[:8]:
        compact = re.sub(r"\s+", "", line)
        if any(label in compact for label in ignored) or re.search(r"[-−]?\d[\d,]*\.\d{2}", compact):
            continue
        if len(re.findall(r"[\u4e00-\u9fff]", compact)) >= 2:
            return compact
    return ""


def extract_known_fields(lines: OCRLines) -> dict[str, str]:
    cleaned = [line.strip() for line in lines if line and line.strip()]
    text = "\n".join(cleaned)
    # Require cents to avoid reading a phone status-bar time such as "12:29" as an amount.
    amount_match = re.search(rf"[-−]?\s*[¥￥]?\s*({AMOUNT_PATTERN})", text)
    transaction_match = re.search(r"(?:交易单号|订单号|交易号|转账单号)\s*[:：]?\s*([A-Za-z0-9]{8,})", text)
    personal_payee, personal_description = _personal_payment_title(cleaned)
    title_merchant = _unlabelled_merchant_title(cleaned)

    return {
        "payment_amount": amount_match.group(1).replace(",", "") if amount_match else "",
        "payment_time": _payment_time(text),
        "merchant_name": _next_value(cleaned, {"商户全称", "收款方", "商家名称", "商户"}) or personal_payee or title_merchant,
        "product_name": _next_value(cleaned, {"商品", "商品名称", "订单内容"}) or personal_description or title_merchant,
        "transaction_number": transaction_match.group(1) if transaction_match else "",
    }


def extract_invoice_fields(lines: OCRLines) -> dict[str, str]:
    cleaned = [line.strip() for line in lines if line and line.strip()]
    text = "\n".join(cleaned)
    # OCR may put a label and its value on adjacent lines; use a compact view
    # only for the labelled-value fallback, while retaining line boundaries.
    compact = re.sub(r"\s+", "", text)
    currency_amounts = re.findall(rf"[¥￥]\s*({AMOUNT_PATTERN})", text)
    # Prefer the explicitly labelled tax-inclusive total.  The tax amount is
    # often printed after the item amount and must never be selected as total.
    amount_match = re.search(rf"(?:价税合计|价税合计（小写）|小写)\s*[:：]?\s*[¥￥]?\s*({AMOUNT_PATTERN})", text)
    if not amount_match:
        amount_match = re.search(rf"(?:价税合计|小写)[:：]?[^\d]{{0,30}}({AMOUNT_PATTERN})", compact)
    if not amount_match:
        amount_match = re.search(rf"(?<!税)(?:合计|金额)\s*[:：]?\s*[¥￥]?\s*({AMOUNT_PATTERN})", text)
    if not amount_match and currency_amounts:
        candidates = [float(value.replace(',', '')) for value in currency_amounts]
        amount_match = re.search(rf"({AMOUNT_PATTERN})", f"{max(candidates):.2f}")
    number_match = re.search(r"(?:发票号码|发票号)\s*[:：]?\s*([A-Za-z0-9]{6,})", text)
    invoice_number = number_match.group(1) if number_match else next((line for line in cleaned if re.fullmatch(r"\d{16,24}", line)), "")
    merchant = _invoice_seller(cleaned)
    item_candidates = [
        line for line in cleaned
        if len(line) < 80
        and any(token in line for token in ("餐饮", "住宿", "交通", "服务", "*"))
        and not any(token in line for token in ("更多服务", "更多发票", "发票查验", "国家税务", "数智化平台", "服务平台", "发票管家", "发送到邮箱", "打印二维码", "下载到手机"))
        and not re.search(r"[￥¥]\s*\d", line)
    ]
    invoice_item = next((line for line in item_candidates if "*" in line or "餐饮" in line or "住宿" in line), item_candidates[0] if item_candidates else "")
    return {
        "invoice_amount": amount_match.group(1).replace(",", "") if amount_match else "",
        "invoice_number": invoice_number,
        "invoice_date": _payment_time(text),
        "invoice_merchant": merchant,
        "invoice_item": invoice_item,
    }


def classify_expense(row: dict[str, str]) -> str:
    """Assign a conservative reimbursement category from merchant and item text."""
    text = " ".join(
        row.get(key, "")
        for key in ("商家名称", "收款方", "商户全称", "商品", "商品名称", "备注")
    ).lower()
    categories = (
        ("住宿费", ("酒店", "宾馆", "旅馆", "民宿", "住宿")),
        ("交通费", ("滴滴", "高德", "打车", "出租", "地铁", "公交", "火车", "动车", "高铁", "机票", "航班", "停车", "加油")),
        ("工具/材料费", ("工具", "材料", "五金", "办公", "文具", "设备", "配件", "耗材", "采购", "商贸", "建材", "电器")),
        ("餐费", ("餐", "饭", "食堂", "餐饮", "美团", "饿了么", "咖啡", "奶茶", "便利店")),
    )
    return next((name for name, words in categories if any(word in text for word in words)), "其他")


def _date_text(value: str) -> tuple[int, int, int] | None:
    matched = re.search(r"(20\d{2})[-./年]\s*(\d{1,2})[-./月]\s*(\d{1,2})", value)
    if not matched:
        return None
    return tuple(map(int, matched.groups()))


def reimbursement_period(rows: Sequence[dict[str, str]]) -> str:
    """Use invoice dates for the title; payment dates are the fallback when no invoice exists."""
    dates = [date for row in rows for date in [_date_text(row.get("付款时间", ""))] if date]
    if not dates:
        dates = [date for row in rows for date in [_date_text(row.get("发票日期", "") or row.get("_invoice_date", ""))] if date]
    if not dates:
        return ""
    first, last = min(dates), max(dates)
    format_date = lambda date: f"{date[0]:04d}.{date[1]:02d}.{date[2]:02d}"
    return f"{format_date(first)}-{format_date(last)}"


def is_invoice(lines: OCRLines) -> bool:
    text = re.sub(r"\s+", "", "\n".join(lines))
    markers = sum(marker in text for marker in ("发票", "电子发票", "增值税", "价税合计", "税额", "开票日期", "发票号码"))
    long_number = bool(re.search(r"\d{8,24}", text))
    return markers >= 1 and (markers >= 2 or long_number)


COLUMN_ALIASES = {
    "付款金额": "payment_amount",
    "支付金额": "payment_amount",
    "交易金额": "payment_amount",
    "金额": "payment_amount",
    "付款时间": "payment_time",
    "支付时间": "payment_time",
    "交易时间": "payment_time",
    "商家名称": "merchant_name",
    "商户名称": "merchant_name",
    "商户全称": "merchant_name",
    "收款方": "merchant_name",
    "商品": "product_name",
    "商品名称": "product_name",
    "备注": "product_name",
    "交易单号": "transaction_number",
    "订单号": "transaction_number",
    "发票金额": "invoice_amount",
    "发票号码": "invoice_number",
    "发票号": "invoice_number",
    "发票日期": "invoice_date",
}


def build_row(columns: Sequence[str], lines: OCRLines, source_name: str) -> dict[str, str]:
    fields = extract_known_fields(lines)
    row = {"源文件": source_name}
    alias_index = {normalize_column_name(name): value for name, value in COLUMN_ALIASES.items()}
    for column in columns:
        key = alias_index.get(normalize_column_name(column))
        row[column] = fields.get(key, "") if key else ""
    return row


def build_payment_rows(columns: Sequence[str], documents: Sequence[tuple[str, OCRLines]]) -> tuple[list[str], list[dict[str, str]]]:
    """Turns a mixed batch into one row per payment, with optional matched invoice fields."""
    result_columns = list(columns)
    if "是否有发票" not in result_columns:
        result_columns.append("是否有发票")

    invoices = [{**extract_invoice_fields(lines), "_source": name} for name, lines in documents if is_invoice(lines)]
    rows: list[dict[str, str]] = []
    matched_invoice_sources: set[str] = set()
    for source_name, lines in documents:
        if is_invoice(lines):
            continue
        row = build_row(result_columns, lines, source_name)
        payment_amount = row.get("付款金额") or row.get("支付金额") or row.get("交易金额") or row.get("金额") or ""
        payment_date = _date_text(row.get("付款时间", ""))
        # Personal QR receipts may put the real merchant in the remark/product
        # field while the merchant field only contains an account display name
        # such as ``商户_陈岚``.  Match against all user-visible identity fields.
        payment_merchant = "".join(
            (row.get(key) or "").replace(" ", "")
            for key in ("商家名称", "收款方", "商品", "商品名称", "备注")
        )
        def is_match(invoice: dict[str, str]) -> bool:
            if not payment_amount or invoice["invoice_amount"] != payment_amount:
                return False
            invoice_date = _date_text(invoice.get("invoice_date", ""))
            invoice_merchant = _usable_invoice_merchant(invoice.get("invoice_merchant", ""))
            date_known = bool(payment_date and invoice_date)
            merchant_known = bool(payment_merchant and invoice_merchant)
            date_ok = date_known and payment_date == invoice_date
            payment_bigrams = {payment_merchant[i:i + 2] for i in range(len(payment_merchant) - 1)}
            invoice_bigrams = {invoice_merchant[i:i + 2] for i in range(len(invoice_merchant) - 1)}
            meaningful_common = (payment_bigrams & invoice_bigrams) - {"公司", "服务", "餐饮", "孝昌"}
            merchant_ok = merchant_known and (
                payment_merchant in invoice_merchant
                or invoice_merchant in payment_merchant
                or SequenceMatcher(None, payment_merchant, invoice_merchant).ratio() >= 0.28
                or bool(meaningful_common)
            )
            if date_known and merchant_known:
                return date_ok and merchant_ok
            return date_ok or merchant_ok
        matched = next((invoice for invoice in invoices if is_match(invoice)), None)
        if matched:
            for column in result_columns:
                alias = COLUMN_ALIASES.get(column)
                if alias in matched:
                    row[column] = matched[alias]
            row["是否有发票"] = "有（金额匹配）"
            row["_invoice_source"] = matched["_source"]
            matched_invoice_sources.add(matched["_source"])
            if matched["invoice_date"]:
                row["_invoice_date"] = matched["invoice_date"]
                payment_date = _date_text(payment_amount)  # populated below by the explicit payment-time check
                payment_date = _date_text(row.get("付款时间", ""))
                invoice_date = _date_text(matched["invoice_date"])
                if payment_date and invoice_date:
                    try:
                        delta = abs((datetime(*payment_date) - datetime(*invoice_date)).days)
                        if delta > 90 or payment_date[0] != invoice_date[0]:
                            row["_invoice_date_warning"] = "日期需核对"
                    except ValueError:
                        row["_invoice_date_warning"] = "日期需核对"
        else:
            row["是否有发票"] = "无"
        rows.append(row)

    # Optional standalone reimbursement: an invoice with no trustworthy
    # payment match is still a reimbursable item, but is explicitly marked so
    # it cannot be mistaken for a payment-backed transaction.
    for invoice in invoices:
        source = invoice.get("_source", "")
        if source in matched_invoice_sources or not invoice.get("invoice_amount"):
            continue
        row = {"源文件": source}
        for column in result_columns:
            row[column] = ""
        amount = invoice.get("invoice_amount", "")
        date = invoice.get("invoice_date", "")
        merchant = invoice.get("invoice_merchant", "")
        row["_invoice_amount"] = amount
        row["_invoice_date"] = date
        row["_invoice_merchant"] = merchant
        row["_invoice_number"] = invoice.get("invoice_number", "")
        row["_invoice_item"] = invoice.get("invoice_item", "")
        row["是否有发票"] = "仅发票（无支付记录）"
        row["_invoice_source"] = source
        row["_invoice_only"] = "1"
        rows.append(row)
    return result_columns, rows


def create_workbook(columns: Sequence[str], rows: Sequence[dict[str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "提取结果"
    headers = ["源文件", *columns]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        sheet.column_dimensions[letter].width = min(40, max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _add_compact_image(sheet, cell: str, image_bytes: bytes | None, width: int, height: int) -> bool:
    """Bad or unsupported uploads should not prevent the user from exporting the text data."""
    if not image_bytes:
        return False
    try:
        if image_bytes.startswith(b"%PDF"):
            import pymupdf

            document = pymupdf.open(stream=image_bytes, filetype="pdf")
            image_bytes = document[0].get_pixmap(matrix=pymupdf.Matrix(1.2, 1.2), alpha=False).tobytes("png")
            document.close()
        image = ExcelImage(BytesIO(image_bytes))
    except Exception:
        return False
    image.width = width
    image.height = height
    sheet.add_image(image, cell)
    return True

def _lookup_image(images: dict[str, bytes], source: str) -> bytes | None:
    """Resolve images even when a persisted history row normalized the filename."""
    if source in images:
        return images[source]
    source_name = Path(source).name
    for name, content in images.items():
        if Path(name).name == source_name:
            return content
    return None


def create_reimbursement_workbook(rows: Sequence[dict[str, str]], payment_images: dict[str, bytes], invoice_images: dict[str, bytes], invoice_rows: Sequence[dict[str, str]] | None = None) -> bytes:
    """Writes a compact, filled-only reimbursement workbook based on the supplied layout."""
    template_path = Path(__file__).resolve().parents[2] / "templates" / "reimbursement-template.xlsx"
    # All reimbursable items are summarized in payment detail.  Invoice-only
    # rows remain visibly distinct through their status and source fields.
    payment_rows = list(rows)
    standalone_rows = []
    workbook = load_workbook(template_path)
    source_sheet = workbook["支付明细"]
    payment_sheet = workbook.create_sheet("支付明细", 0)
    payment_sheet.sheet_view.showGridLines = False
    payment_sheet.freeze_panes = "A3"
    for letter in "ABCDEFGH":
        payment_sheet.column_dimensions[letter].width = source_sheet.column_dimensions[letter].width
    payment_sheet.column_dimensions["H"].width = 12
    for column in range(1, 9):
        for row_number in (1, 2):
            source = source_sheet.cell(row_number, column)
            target = payment_sheet.cell(row_number, column, source.value)
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
    payment_sheet.merge_cells("A1:H1")
    payment_sheet.row_dimensions[1].height = max(36, source_sheet.row_dimensions[1].height or 36)
    payment_sheet.row_dimensions[2].height = max(28, source_sheet.row_dimensions[2].height or 28)
    prefix = str(source_sheet["A1"].value or "刘生费用报销单").split("（", 1)[0].split("(", 1)[0]
    period = reimbursement_period(rows)
    payment_sheet["A1"] = f"{prefix}（{period}）" if period else prefix

    for index, row in enumerate(payment_rows, start=3):
        invoice_only = bool(row.get("_invoice_only"))
        amount_text = row.get("付款金额") or row.get("支付金额") or row.get("交易金额") or row.get("金额") or (row.get("_invoice_amount") if invoice_only else "") or "0"
        try:
            amount = float(amount_text.replace(",", ""))
        except ValueError:
            amount = 0
        date_time = row.get("付款时间", "") or (row.get("_invoice_date", "") if invoice_only else "")
        date_value, _, time_value = date_time.partition(" ")
        invoice_status = row.get("是否有发票", "")
        merchant = row.get("商家名称") or row.get("收款方") or (row.get("_invoice_merchant") if invoice_only else "") or ""
        expense = row.get("_invoice_item") if invoice_only else ""
        values = [index - 2, date_value, time_value, merchant, amount, expense or classify_expense(row), "仅发票" if invoice_only else ("有票" if invoice_status.startswith("有") else "无票")]
        for column, value in enumerate(values, start=1):
            source = source_sheet.cell(3, column)
            target = payment_sheet.cell(index, column, value)
            target._style = copy(source._style)
            target.alignment = copy(source.alignment)
            target.border = copy(source.border)
        payment_sheet.cell(index, 5).number_format = "0.00"
        edge = Side(style="thin", color="B7B7B7")
        for column in (7, 8):
            payment_sheet.cell(index, column).border = Border(left=edge, right=edge, top=edge, bottom=edge)
        image_bytes = _lookup_image(payment_images, row.get("源文件", ""))
        _add_compact_image(payment_sheet, f"H{index}", image_bytes, 84, 150)
        payment_sheet.row_dimensions[index].height = 112
        payment_sheet.cell(index, 7).fill = PatternFill("solid", fgColor="E2F0D9" if values[6] == "有票" else ("FFF2CC" if values[6] == "仅发票" else "FCE4D6"))

    total_row = len(payment_rows) + 3
    for column in range(1, 9):
        source = source_sheet.cell(193, column)
        target = payment_sheet.cell(total_row, column)
        target._style = copy(source._style)
        target.alignment = copy(source.alignment)
        target.border = copy(source.border)
    payment_sheet.cell(total_row, 1, "合计")
    payment_sheet.cell(total_row, 5, f"=SUM(E3:E{total_row - 1})" if rows else 0)
    payment_sheet.cell(total_row, 5).number_format = "0.00"
    payment_sheet.row_dimensions[total_row].height = max(28, source_sheet.row_dimensions[193].height or 28)
    edge = Side(style="thin", color="B7B7B7")
    for column in range(1, 9):
        payment_sheet.cell(total_row, column).border = Border(left=edge, right=edge, top=edge, bottom=edge)
    payment_sheet.auto_filter.ref = f"A2:H{total_row - 1}" if payment_rows else "A2:H2"
    payment_sheet.print_area = f"A1:H{total_row}"

    # Remove the completed-example tabs and build a clean invoice tab containing only matched invoices.
    for sheet in list(workbook.worksheets):
        if sheet is not payment_sheet:
            workbook.remove(sheet)
    payment_sheet.title = "支付明细"
    invoice_sheet = workbook.create_sheet("发票明细")
    invoice_sheet.sheet_view.showGridLines = False
    invoice_sheet.merge_cells("A1:G1")
    invoice_sheet["A1"] = payment_sheet["A1"].value
    invoice_sheet["A1"]._style = copy(payment_sheet["A1"]._style)
    invoice_sheet.row_dimensions[1].height = payment_sheet.row_dimensions[1].height
    invoice_sheet.append(["序号", "支付金额", "发票金额", "发票号码", "是否匹配", "日期核对", "发票图片"])
    invoice_sheet.freeze_panes = "A3"
    table_edge = Side(style="thin", color="B7B7B7")
    invoice_sheet.row_dimensions[2].height = max(28, payment_sheet.row_dimensions[2].height or 28)
    for cell in invoice_sheet[2]:
        cell.font = Font(bold=True)
        cell.border = Border(left=table_edge, right=table_edge, top=table_edge, bottom=table_edge)
        cell.alignment = copy(payment_sheet["A2"].alignment)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for letter, width in {"A": 8, "B": 13, "C": 13, "D": 18, "E": 14, "F": 14, "G": 22}.items():
        invoice_sheet.column_dimensions[letter].width = width
    matched_payments = {row.get("_invoice_source", ""): row for row in payment_rows if row.get("_invoice_source") and not row.get("_invoice_only")}
    if invoice_rows is None:
        invoice_rows = [{
            "_source": source,
            "invoice_amount": payment.get("发票金额", ""),
            "invoice_number": payment.get("发票号码", ""),
        } for source, payment in matched_payments.items()]
    invoice_index = 3
    for invoice in invoice_rows:
        source = invoice.get("_source", "")
        payment = matched_payments.get(source)
        invoice_sheet.append([
            invoice_index - 2,
            payment.get("付款金额", "") if payment else "",
            invoice.get("invoice_amount", ""),
            invoice.get("invoice_number", ""),
            "金额匹配" if payment else "仅发票（无支付记录）",
            payment.get("_invoice_date_warning", "正常") if payment else "需核对",
        ])
        image_bytes = _lookup_image(invoice_images, source) or _lookup_image(payment_images, source)
        _add_compact_image(invoice_sheet, f"G{invoice_index}", image_bytes, 150, 125)
        for cell in invoice_sheet[invoice_index]:
            cell.border = Border(left=table_edge, right=table_edge, top=table_edge, bottom=table_edge)
            cell.alignment = copy(payment_sheet["A3"].alignment)
            cell.fill = PatternFill("solid", fgColor="F7FBFF" if invoice_index % 2 == 0 else "FFFFFF")
        invoice_sheet.row_dimensions[invoice_index].height = 94
        invoice_index += 1
    if invoice_index == 3:
        workbook.remove(invoice_sheet)
    if standalone_rows:
        standalone = workbook.create_sheet("发票单独报销")
        standalone.sheet_view.showGridLines = False
        standalone.merge_cells("A1:G1")
        standalone["A1"] = payment_sheet["A1"].value
        standalone["A1"]._style = copy(payment_sheet["A1"]._style)
        standalone.append(["序号", "开票日期", "销售方", "费用用途", "报销金额", "发票号码", "发票图片"])
        for cell in standalone[2]:
            cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="FCE4D6")
            cell.border = Border(left=table_edge, right=table_edge, top=table_edge, bottom=table_edge)
            cell.alignment = copy(payment_sheet["A2"].alignment)
        for letter, width in {"A": 8, "B": 14, "C": 30, "D": 24, "E": 14, "F": 22, "G": 22}.items():
            standalone.column_dimensions[letter].width = width
        for index, row in enumerate(standalone_rows, start=3):
            standalone.append([index - 2, row.get("_invoice_date", ""), row.get("_invoice_merchant", ""), row.get("_invoice_item", "") or "发票单独报销", float(row.get("_invoice_amount", "0") or 0), row.get("_invoice_number", "")])
            image_bytes = _lookup_image(invoice_images, row.get("源文件", "")) or _lookup_image(payment_images, row.get("源文件", ""))
            _add_compact_image(standalone, f"G{index}", image_bytes, 150, 125)
            for cell in standalone[index]:
                cell.border = Border(left=table_edge, right=table_edge, top=table_edge, bottom=table_edge)
                cell.alignment = copy(payment_sheet["A3"].alignment)
            standalone.cell(index, 5).number_format = "0.00"
            standalone.row_dimensions[index].height = 94
        total = len(standalone_rows) + 3
        standalone.cell(total, 1, "合计")
        standalone.cell(total, 5, f"=SUM(E3:E{total - 1})")
        standalone.cell(total, 5).number_format = "0.00"
        for cell in standalone[total]:
            cell.border = Border(left=table_edge, right=table_edge, top=table_edge, bottom=table_edge)
    output = BytesIO(); workbook.save(output); return output.getvalue()


def reimbursement_workbook_title(rows: Sequence[dict[str, str]]) -> str:
    """Return the same title text written into the workbook's first row."""
    template_path = Path(__file__).resolve().parents[2] / "templates" / "reimbursement-template.xlsx"
    workbook = load_workbook(template_path, read_only=True)
    raw = str(workbook["支付明细"]["A1"].value or "刘生费用报销单")
    prefix = raw.split("（", 1)[0].split("(", 1)[0]
    period = reimbursement_period(rows)
    return f"{prefix}（{period}）" if period else prefix


class LocalReceiptExtractor:
    """Reads images locally; OCR model weights are loaded on the host, never sent to an API."""

    def __init__(self) -> None:
        self.engine = getenv("RECEIPT_OCR_ENGINE", "auto").lower()
        self._paddle = None
        if self.engine in {"paddle", "auto"}:
            try:
                from paddleocr import PaddleOCR

                # PaddleOCR 3.x API.  Disable orientation/unwarping here because
                # payment screenshots are already upright and CPU cost matters.
                self._paddle = PaddleOCR(
                    lang="ch",
                    use_doc_orientation_classify=False,
                    use_doc_unwarping=False,
                    use_textline_orientation=False,
                )
                self.engine = "paddle"
            except Exception:
                if self.engine == "paddle":
                    raise
                self.engine = "rapidocr"

        from rapidocr_onnxruntime import RapidOCR

        # Keep the server usable during a batch: two OCR sessions with two CPU threads each.
        # Both values can be adjusted at deployment time without a code change.
        thread_count = int(getenv("RECEIPT_OCR_THREADS", "2"))
        self._ocr = RapidOCR(intra_op_num_threads=thread_count, inter_op_num_threads=1)

    def _read_paddle(self, image_bytes: bytes) -> list[str]:
        """Return PaddleOCR text as the same line interface used by the matcher."""
        if not self._paddle:
            return []
        try:
            from PIL import Image
            import numpy as np
            image = np.asarray(Image.open(BytesIO(image_bytes)).convert("RGB"))
            result = self._paddle.predict(image)
            lines: list[str] = []
            for item in result or []:
                data = item.json if hasattr(item, "json") else item
                if callable(data):
                    data = data()
                if isinstance(data, str):
                    import json
                    data = json.loads(data)
                if not isinstance(data, dict):
                    continue
                payload = data.get("res", data)
                texts = payload.get("rec_texts", []) if isinstance(payload, dict) else []
                lines.extend(str(value).strip() for value in texts if str(value).strip())
            return lines
        except Exception:
            return []

    def read(self, image_bytes: bytes) -> list[str]:
        if image_bytes.startswith(b"%PDF"):
            from pypdf import PdfReader

            reader = PdfReader(BytesIO(image_bytes))
            text_lines = [line.strip() for page in reader.pages for line in (page.extract_text() or "").splitlines() if line.strip()]
            if text_lines:
                return text_lines
            # Scanned/image-only PDFs have no text layer; render the first page
            # and send it through the same local OCR pipeline as screenshots.
            try:
                import pymupdf
                document = pymupdf.open(stream=image_bytes, filetype="pdf")
                if document.page_count:
                    image_bytes = document[0].get_pixmap(matrix=pymupdf.Matrix(1.8, 1.8), alpha=False).tobytes("png")
                document.close()
            except Exception:
                return []
        try:
            from PIL import Image
            source = Image.open(BytesIO(image_bytes)).convert("RGB")
            if max(source.width, source.height) > 2200:
                scale = 2200 / max(source.width, source.height)
                source = source.resize((int(source.width * scale), int(source.height * scale)), Image.Resampling.LANCZOS)
                prepared = BytesIO(); source.save(prepared, format="JPEG", quality=90)
                image_bytes = prepared.getvalue()
        except Exception:
            pass
        if self.engine == "paddle":
            paddle_lines = self._read_paddle(image_bytes)
            if paddle_lines:
                return paddle_lines
        result, _elapsed = self._ocr(image_bytes)
        lines = [item[1] for item in result or []]
        if lines:
            return lines
        try:
            from PIL import Image, ImageEnhance, ImageOps
            source = Image.open(BytesIO(image_bytes)).convert("RGB")
            scale = max(1.0, min(2.5, 1800 / max(source.width, source.height)))
            if scale > 1:
                source = source.resize((int(source.width * scale), int(source.height * scale)), Image.Resampling.LANCZOS)
            enhanced = ImageEnhance.Contrast(ImageOps.grayscale(source)).enhance(1.8)
            output = BytesIO(); enhanced.save(output, format="PNG")
            retry, _elapsed = self._ocr(output.getvalue())
            return [item[1] for item in retry or []]
        except Exception:
            return lines

    def read_many(self, images: Sequence[bytes], progress: Callable[[int, int, int], None] | None = None, worker_count: int | None = None) -> list[list[str]]:
        """Run a small number of independent local OCR sessions concurrently."""
        if len(images) < 2:
            results = [self.read(image) for image in images]
            if images and progress:
                progress(1, 1, 0)
            return results
        configured_workers = int(getenv("RECEIPT_OCR_WORKERS", "2")) if worker_count is None else worker_count
        worker_count = min(max(1, configured_workers), len(images), 4)
        sessions: Queue[LocalReceiptExtractor] = Queue()
        sessions.put(self)
        for _ in range(worker_count - 1):
            sessions.put(LocalReceiptExtractor())

        def read_one(image: bytes) -> list[str]:
            reader = sessions.get()
            try:
                return reader.read(image)
            finally:
                sessions.put(reader)

        results: list[list[str] | None] = [None] * len(images)
        with ThreadPoolExecutor(max_workers=worker_count, thread_name_prefix="receipt-ocr") as executor:
            futures = {executor.submit(read_one, image): index for index, image in enumerate(images)}
            for completed, future in enumerate(as_completed(futures), start=1):
                index = futures[future]
                try:
                    results[index] = future.result()
                except Exception:
                    results[index] = []
                if progress:
                    progress(completed, len(images), index)
        return [result or [] for result in results]
