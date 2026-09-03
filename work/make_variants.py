import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
src = r"F:\project\nixiang\outlook-batch-manager\work\quote_openpyxl.xlsx"
out_root = r"F:\project\nixiang\outlook-batch-manager\work"


def fresh(path):
    wb = load_workbook(path)
    return wb


# 1. Drop the terms sheet.
wb = fresh(src)
if "报价说明" in wb.sheetnames:
    del wb["报价说明"]
wb.save(rf"{out_root}\v_no_terms.xlsx")

# 2. Remove merged cells.
wb = fresh(src)
for ws in wb.worksheets:
    for rng in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(rng))
        except KeyError:
            pass
wb.save(rf"{out_root}\v_no_merge.xlsx")

# 3. Remove formulas.
wb = fresh(src)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                c.value = None
wb.save(rf"{out_root}\v_no_formulas.xlsx")

# 4. Remove row heights and column widths.
wb = fresh(src)
for ws in wb.worksheets:
    ws.row_dimensions.clear()
    ws.column_dimensions.clear()
wb.save(rf"{out_root}\v_no_sizes.xlsx")

# 5. Keep only first 40 rows of the quote sheet and drop terms.
wb = fresh(src)
if "报价说明" in wb.sheetnames:
    del wb["报价说明"]
ws = wb["报价单"]
for row in ws.iter_rows(min_row=41):
    for c in row:
        c.value = None
ws.row_dimensions.clear()
ws.column_dimensions.clear()
wb.save(rf"{out_root}\v_first_rows.xlsx")

print("variants saved")
