import json
import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

src = r"C:\Users\yiliu\Desktop\20260717-泰国报价单(2)(1).xlsx"
wb = load_workbook(src, data_only=False)
out = {"sheets": []}
for ws in wb.worksheets:
    merged = [str(r) for r in ws.merged_cells.ranges]
    rows = []
    for row in ws.iter_rows():
        vals = []
        for c in row:
            vals.append(
                {
                    "cell": c.coordinate,
                    "v": c.value,
                    "f": c.data_type,
                }
            )
        rows.append(vals)
    out["sheets"].append(
        {
            "name": ws.title,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged": merged,
            "rows": rows,
        }
    )
print(json.dumps(out, ensure_ascii=False, indent=1))
